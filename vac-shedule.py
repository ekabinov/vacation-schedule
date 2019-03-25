# -*- coding: utf-8 -*-

import configparser
import openpyxl
from openpyxl import Workbook


import time
import datetime
from datetime import timedelta
from datetime import datetime
import logging
import os.path

sotr = []
id = []
all_days = []
all_days_all = 0
days_in_year = 365  # сделать функцией определения кол-ва дней в году по году
otpusk_days = []

# TODO: пять отпусков в году, потом расширить
start_day1 = []
start_day2 = []
start_day3 = []
start_day4 = []
start_day5 = []

otp_days1 = []
otp_days2 = []
otp_days3 = []
otp_days4 = []
otp_days5 = []


# TODO: считывать из файла сотрудников: формат: ФИО, дата начала/кол-во дней
# TODO: переделать файл со списком, хранить в xls, json, xml
# TODO: нарисовать календарь в xls (с праздниками и выходными)
# TODO: нарисовать календарь в pdf
# TODO: нарисовать календарь в png
# TODO: анализ пересечения между сотрудниками
# TODO: процент одновременного нахождения сотрудников в отпуске по дням с графиком
# TODO: экспорт в формы для отпусков

# TODO: потом сделать проверку на формат файла try...


def read_from_ini_file(year):
    global all_days
    global all_days_all

    config = configparser.RawConfigParser()
    config.read(str(year)+'.txt')

    if os.path.isfile(str(year)+'.txt'):
        count_sotr = config.getint('options', 'sotr_count')

        for i in range(0, count_sotr):
            sotr.append(config.get('sotrudnik' + str(i + 1), 'name'))
            all_days.append(config.get('sotrudnik' + str(i + 1), 'days_total'))

        for i in range(len(sotr)):
            all_days_all = all_days_all + int(all_days[i])
            # print(sotr[i], ' ', all_days[i])

        print("\nИтого дней на всех:", all_days_all)

    else:
        # print("нет файла", str(year)+'.txt')
        exit(1)


def raschet(sotr_otpusk_days):
    global otpusk_days

    for i in range(1, days_in_year + 1):
        if i == 10 or i == 11 or i == 12 or i == 13:
            sotr_otpusk_days = sotr_otpusk_days + "1"
        else:
            sotr_otpusk_days = sotr_otpusk_days + "0"

    print(sotr_otpusk_days)

"""
def raschet2():
    global otpusk_days

    for j in range(1, 5):
        for i in range(1, days_in_year + 1):
            otpusk_days[j] = otpusk_days[j] + "0"

    print(otpusk_days[1])

raschet2()
"""

if __name__ == '__main__':

    plan_otpusk_year = input("Введите год планирования отпусков:")
    # print("Год планирования отпуска:", plan_otpusk_year, "\n")

    read_from_ini_file(plan_otpusk_year)

    # raschet(otpusk_days[1])

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "test"
    # d = ws.cell(row=4, column=2, value=10)
    wb.save(plan_otpusk_year + '.xlsx')
